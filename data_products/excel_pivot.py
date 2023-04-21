import requests
import os
from openpyxl import Workbook

# Define the URL of the CSV file
def create_excel(csv_url=os.environ.get('CSV_URL')):
    # Fetch the CSV data from the URL
    response = requests.get(csv_url)
    csv_data = response.content.decode('utf-8')

    # Load the CSV data into a new workbook in memory
    wb = Workbook()
    ws = wb.active

    rows = csv_data.split('\n')
    for r in rows:
        ws.append(r.split(','))

    # Pivot the table using the Excel API
    ws_pivot = wb.create_sheet('Pivoted Table')

    # create a dictionary to store the pivot table data
    pivot_table = {}

    # iterate over each row of the worksheet, starting from the second row
    for row in ws.iter_rows(min_row=2, values_only=True):
        # get the plan name, recognized at date, and monthly revenue
        plan_name = row[0]
        recognized_at = row[1]
        monthly_revenue = row[2]

        # check if the plan name is already in the pivot table dictionary
        if plan_name not in pivot_table:
            # if not, add a new dictionary entry for the plan name
            pivot_table[plan_name] = {}

        # set the monthly revenue for the recognized at date and plan name
        pivot_table[plan_name][recognized_at] = monthly_revenue

    # sort the recognized at dates in ascending order
    recognized_at_dates = sorted(list(set([date for plan in pivot_table.values() for date in plan.keys() if date is not None])))

    # create the header row for the pivot table
    header_row = ['plan_name'] + recognized_at_dates

    # create a list to store the rows of the pivot table
    pivot_rows = [header_row]

    # iterate over each plan in the pivot table dictionary
    for plan_name, plan_data in pivot_table.items():
        # create a new row for the pivot table
        row = [plan_name]

        # iterate over each recognized at date in the pivot table
        for recognized_at in recognized_at_dates:
            # get the monthly revenue for the recognized at date and plan name
            monthly_revenue = plan_data.get(recognized_at, '')

            # add the monthly revenue to the row
            row.append(monthly_revenue)

        # add the row to the pivot table
        pivot_rows.append(row)

    # write the pivot table to the worksheet
    for row in pivot_rows:
        ws_pivot.append(row)

    print('TODO: Support dynamic import via URL, do pivot, and create financial models sheet')

    # Save the Excel workbook
    file = 'financial_models.xlsx'
    wb.save(file)
    print(f'Created {file}')


if __name__ == '__main__':
    create_excel()
